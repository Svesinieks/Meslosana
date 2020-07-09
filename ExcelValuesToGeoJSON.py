import json
from openpyxl import load_workbook

wb2 = load_workbook('Book1.xlsx')
ws = wb2.active


with open('LaukiGeojson/Paraugu_dati.geojson', 'r+') as f:
    data = json.load(f)

    for i, feature in enumerate(data['features']):
        j = 2
        while j <= 49:
            id = ws['A'+str(j)]
            if feature['properties']["ID"] == id.value:
                print(feature['properties']["ID"],id.value)
                mapKg = ws['B'+str(j)]
                data['features'][i]['properties']["mapkg/h"]=mapKg.value
                mapPro = ws['C' + str(j)]
                data['features'][i]['properties']["map%"] = mapPro.value
                kclKG = ws['D' + str(j)]
                data['features'][i]['properties']["kclkg/ha"] = kclKG.value
                kclPro = ws['E' + str(j)]
                data['features'][i]['properties']["kcl%"] = kclPro.value
                catHa = ws['F' + str(j)]
                data['features'][i]['properties']["cat/ha"] = catHa.value
                caPro = ws['G' + str(j)]
                data['features'][i]['properties']["ca%"] = caPro.value
                f.seek(0)
                json.dump(data, f, indent=4)
                f.truncate()
            j +=1
print(feature['properties'])


